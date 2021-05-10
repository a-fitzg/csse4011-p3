from openpyxl import load_workbook
import re
import tago

MY_DEVICE_TOKEN = '3503290b-05e5-433d-a864-e1b8e7bfbf11'
my_device = tago.Device(MY_DEVICE_TOKEN)

Y_train = []
X_train = []
ds = []

wb = load_workbook(filename=r'C:\Users\desmo\csse4011\csse4011-p3\TrainingData.xlsx')
sheet = wb.active
for r in sheet.iter_rows(min_row=1,max_row=4500,min_col=1,max_col=2,values_only=True):
    ds.append(r)
 
print(ds)   
rssi= {
        'variable': 'trainingset',
        'value': ds
}


result1 = my_device.insert(rssi)  # With response
if result1['status']:
  print(result1['result'])
else:
  print(result1['message'])

# for r in sheet.iter_rows(min_row=1,max_row=4500,min_col=1,max_col=2,values_only=True):
#     r_n = [int(s) for s in r[0].split(',')]
#     r_yn = [int(s) for s in r[1].split(',')]
   
#     X_train.append(r_n)
#     Y_train.append(r_yn)

# rssi= {
#         'variable': 'training_rssi',
#         'value': X_train
#     }
# loc = {
#         'variable': 'training_loc',
#         'value': Y_train
# }
    

# result1 = my_device.insert(rssi)  # With response
# if result1['status']:
#   print(result1['result'])
# else:
#   print(result1['message'])
# result2 = my_device.insert(loc)  # With response

# if result2['status']:
#   print(result2['result']) 
# else:
#   print(result2['message'])

